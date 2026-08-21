"""
File serving endpoints for generated attachments.

GET  /api/files/{message_id}/{attachment_id}
GET  /api/files/{message_id}/{attachment_id}/preview?format=html

Auth: JWT via existing get_current_user dependency. The endpoint verifies
that the requested attachment belongs to a conversation owned by the
caller — cross-user access returns 404 (not 403, to avoid leaking
attachment existence).

Caching: an attachment id is a fresh uuid4 minted by
`doc_generators._persist`, and nothing ever rewrites the storage key it
names. The bytes behind one of these URLs are therefore IMMUTABLE for the
life of the row, which is what lets both routes send a strong `ETag` plus
`max-age=31536000, immutable` and answer a conditional GET with a bodyless
304. See `attachment_etag` for why the validator needs no read.
"""

from __future__ import annotations

import json
import logging
from typing import Optional

import httpx
from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse, HTMLResponse, JSONResponse, Response
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.api.auth import get_current_user, SSO_COOKIE_NAME
from app.config import settings
from app.db import get_db
from app.db.models import Conversation, Message, User
from app.services import decode_access_token, get_user_by_id
from app.services.file_storage import get_storage_backend, stream_file

logger = logging.getLogger(__name__)

router = APIRouter(prefix="/files", tags=["Files"])


# ── Cache validators ────────────────────────────────────────────────────
#
# The inline card was re-downloading the same picture on every visit to a
# thread. `private, max-age=3600` and no validator means the browser has
# exactly two states for a 2.4 MB PNG: fresh for an hour, then a full
# unconditional re-fetch. There was nothing in between, because there was
# no ETag to revalidate WITH — so an image looked at on Monday was pulled
# down again on Tuesday, and the card showed a spinner for it.
#
# A chat attachment cannot change, so `immutable` is not an optimism here,
# it is the fact: `_persist` mints a new uuid4 per file and writes it once.

_ATTACHMENT_CACHE = "private, max-age=31536000, immutable"

# A DERIVED rendering (DOCX/PPTX → PDF via LibreOffice) is deterministic in
# its source but not in the renderer, so it revalidates hourly rather than
# being pinned for a year. The ETag still turns that revalidation into a
# 304 instead of a re-transfer.
_DERIVED_CACHE = "private, max-age=3600"


def attachment_etag(attachment_id: str, size: int, variant: str = "") -> str:
    """Strong validator for one attachment's bytes — computed, never read.

    The id IS the content address: `_persist` mints a uuid4 per file, uses it
    in the storage key, and nothing rewrites that key. So the pair (id, size)
    names exactly one immutable blob and no hashing is needed to prove it —
    which matters, because hashing would mean reading the whole file on a
    request whose entire point is to transfer nothing.

    `variant` separates the inline WebP derivative and the LibreOffice PDF
    render from the original: they share an id but not their bytes, and one
    ETag across all three would let a client hold a thumbnail and be told its
    copy of the full-size image is current.
    """
    tag = f"{attachment_id}-{int(size or 0)}"
    if variant:
        tag = f"{tag}-{variant}"
    return f'"{tag}"'


def _suffix_etag(etag: Optional[str], suffix: str) -> Optional[str]:
    """`"abc-12"` + `pdf` → `"abc-12-pdf"`. Names a DERIVATION of the bytes
    `etag` validates, so a client holding the render and a client holding the
    source are never told the same thing."""
    if not etag:
        return None
    body = etag[2:] if etag.startswith("W/") else etag
    return f'{body[:-1]}-{suffix}"' if body.endswith('"') else f'"{body}-{suffix}"'


def etag_matches(request: Optional[Request], etag: str) -> bool:
    """True when the caller already holds `etag` (RFC 9110 If-None-Match).

    Handles the comma-separated list, `*`, and the `W/` prefix — a GET uses
    the WEAK comparison function, so `W/"x"` and `"x"` match each other.
    """
    if request is None or not etag:
        return False
    header = request.headers.get("if-none-match") or ""
    if not header:
        return False
    if header.strip() == "*":
        return True
    want = etag[2:] if etag.startswith("W/") else etag
    for candidate in header.split(","):
        c = candidate.strip()
        if c.startswith("W/"):
            c = c[2:]
        if c and c == want:
            return True
    return False


def not_modified(etag: str, cache_control: str) -> Response:
    """A 304 carrying the validators and NOTHING else.

    Starlette omits content-length and content-type for 304 on its own; a
    body here would be a protocol violation, and clients that do not simply
    ignore it tend to hang waiting for bytes the framing already ended.
    """
    return Response(status_code=304, headers={"ETag": etag, "Cache-Control": cache_control})


def _conditional_headers(request: Optional[Request]) -> dict:
    """The caller's cache validators, to forward upstream.

    Without this the platform proxy strips every conditional a browser
    sends, so the agent can never answer 304 and the platform — where every
    real user is — re-transfers the file on each revalidation no matter what
    the agent's headers say.
    """
    if request is None:
        return {}
    out = {}
    for h in ("if-none-match", "if-modified-since"):
        v = request.headers.get(h)
        if v:
            out[h] = v
    return out


async def _get_agent_proxy_info(user_id: str, db: AsyncSession):
    """Return (agent_url, agent_api_key) for the caller's remote agent, or None.

    On the platform run_mode, attachments + file bytes live on the VPS agent.
    Platform forwards GET /api/files/* to the agent via X-Agent-Key, mirroring
    the stats / dashboard proxy patterns.
    """
    # NOTE the failure modes below are deliberately distinguishable.
    #
    # This function used to end `except Exception: pass` / `return None`,
    # and the caller turns None into 404 "No active agent for user". So an
    # import error, a DB blip, a decrypt failure and a genuinely absent
    # agent all produced the SAME 404 — a user with a healthy, running
    # agent was told the file did not exist, and nothing was logged.
    #
    # That cost a long hunt on 2026-08-13: the agent was measured serving
    # the exact requested file as a valid PDF in 4.7 s while the platform
    # answered 404 in 579 ms, i.e. without ever asking it. Every layer
    # looked healthy because the only layer that had failed said nothing.
    try:
        from app.db.models import AgentConfig  # noqa — may not exist on agent DBs
        result = await db.execute(
            select(AgentConfig.agent_url, AgentConfig.agent_api_key).where(
                AgentConfig.user_id == user_id,
                AgentConfig.deploy_status == "active",
            )
        )
        row = result.first()
    except Exception:
        # An error here is NOT "you have no agent" — it is our fault, and
        # it must be visible. Still returns None (the caller's contract is
        # unchanged and a file read must not 500), but never silently.
        logger.exception(
            "[files] agent proxy lookup FAILED for user=%s — "
            "the caller will report 404 'no active agent', which is wrong",
            user_id,
        )
        return None

    if row is None:
        logger.warning(
            "[files] no AgentConfig row with deploy_status='active' for "
            "user=%s — file proxying is impossible until one exists", user_id,
        )
        return None
    if not row.agent_url or not row.agent_api_key:
        logger.warning(
            "[files] AgentConfig for user=%s is active but incomplete "
            "(url=%s, key=%s) — cannot proxy",
            user_id, bool(row.agent_url), bool(row.agent_api_key),
        )
        return None
    return (row.agent_url, row.agent_api_key)


# Read timeout for a plain byte-for-byte file proxy. The agent is already
# holding the bytes; anything slower than this is a dead socket.
_PROXY_READ_S = 15.0

# ...but a DOCX/PPTX preview is not a file read. The agent shells out to
# LibreOffice and converts to PDF before the first byte exists, and that
# conversion is measured in SECONDS, not milliseconds:
#
#   cached                    0.2 s
#   warm convert              3.3 s
#   cold convert             10.3 s   <- measured 2026-08-13, real user file
#
# Against the 15 s read budget that is a 1.45x margin on a path whose own
# docstring advertises "~3-8s (LibreOffice cold start)". Any load on the
# agent pushes a first-view preview over the line, and the client sees a
# dead request it renders as "Missing PDF <url>" — which names the file
# and so reads as corruption rather than a timeout. Every generated
# document is a cold convert exactly once, so this lands on first view,
# which is the only view that matters.
_PREVIEW_READ_S = 90.0


async def _proxy_file(
    agent_url: str,
    agent_api_key: str,
    path: str,
    params: dict,
    accept: Optional[str] = None,
    *,
    read_timeout: float = _PROXY_READ_S,
    conditional: Optional[dict] = None,
):
    """Open a streaming GET to the agent's files endpoint and return an
    (content_iter, headers, status) tuple. Caller wraps into StreamingResponse.
    A `status` of 304 means the iterator is empty and the caller must answer
    with a bodyless `Response(304, headers=…)`.

    Uses aiter_bytes (decoded) and does NOT forward content-length — any
    intermediate gzip would make that value wrong, and a stale content-length
    makes the browser wait on bytes that will never arrive (this showed up
    as ~60s downloads for 37 KB files). FastAPI emits chunked transfer
    encoding when content-length is absent, which matches the iterator.

    ``accept`` is the caller's Accept header, forwarded so the agent's
    preview route can tell a browser/WebView navigation (wants an HTML
    fallback page) from a fetch() (wants JSON). Without it httpx sends
    `*/*` and every client would get JSON.

    ``conditional`` carries the caller's If-None-Match / If-Modified-Since.
    A proxy that drops those turns every revalidation into a full transfer,
    which is what made caching invisible in production: the agent has spoken
    Cache-Control since the file routes existed, but on run_mode=platform —
    where every real user is — no validator ever crossed this hop.
    """
    url = f"{agent_url.rstrip('/')}/api/files/{path}"
    # TKT-LAT-007 (wave 3): shared agent_http client. Streaming response
    # is independent of the client lifecycle — we close the response,
    # not the client (which is shared across call sites).
    from app.services.agent_http import get_agent_http_client

    client = get_agent_http_client()
    headers = {"X-Agent-Key": agent_api_key}
    if accept:
        headers["Accept"] = accept
    headers.update(conditional or {})
    req = client.build_request(
        "GET", url,
        headers=headers,
        params=params,
        timeout=httpx.Timeout(connect=5.0, read=read_timeout, write=5.0, pool=5.0),
    )
    resp = await client.send(req, stream=True)
    if resp.status_code >= 400:
        # Pass the agent's error through VERBATIM — status, content-type and
        # body. This used to re-raise as HTTPException(detail=<body text>),
        # which took the agent's `{"detail":"No preview for application/pdf"}`
        # and wrapped it once more into
        # `{"detail":"{\"detail\":\"No preview for application/pdf\"}"}` —
        # the double-encoded blob the iOS WebView painted on a black screen
        # (2026-08-18). The agent already speaks the right shape (JSON for
        # API callers, an HTML fallback page for a browser/WebView asking
        # for text/html); the platform's only job is to relay it.
        body = await resp.aread()
        await resp.aclose()
        err_headers = {}
        for h in ("content-type", "cache-control"):
            v = resp.headers.get(h)
            if v:
                err_headers[h] = v

        async def _one():
            yield body

        return _one(), err_headers, resp.status_code

    if resp.status_code == 304:
        # The agent says the caller's copy is current. There is no body to
        # stream, so close the upstream response here — nobody will iterate
        # it — and relay only the validators.
        await resp.aclose()
        nm_headers = {}
        for h in ("etag", "cache-control", "vary", "last-modified"):
            v = resp.headers.get(h)
            if v:
                nm_headers[h] = v

        async def _none():
            return
            yield b""  # pragma: no cover — makes this a generator

        return _none(), nm_headers, 304

    async def _iter():
        try:
            async for chunk in resp.aiter_bytes(chunk_size=64 * 1024):
                yield chunk
        finally:
            await resp.aclose()

    # Only forward content-type and content-disposition. Let chunked transfer
    # encoding handle framing so the browser doesn't wait on a size that may
    # not match (e.g. after transparent gzip decoding in the httpx layer).
    # `etag` / `last-modified` ride along because a validator the proxy drops
    # is a validator the browser can never send back.
    passthrough_headers = {}
    for h in ("content-type", "content-disposition", "cache-control",
              "etag", "last-modified", "vary"):
        v = resp.headers.get(h)
        if v:
            passthrough_headers[h] = v
    return _iter(), passthrough_headers, resp.status_code


async def _get_user_for_file(
    request: Request,
    token: Optional[str],
    db: AsyncSession,
):
    """Auth for file endpoints, in EXPLICIT-BEFORE-AMBIENT order:
      1) Authorization: Bearer <token> header (standard)
      2) ?token=<jwt> query param — required for <iframe src> / <img src>
         embeds that cannot set headers.
      3) hex_sso_token cookie (SSO, same-origin)
      4) X-Agent-Key header (platform proxy mode, reused from auth.py).

    The cookie is LAST of the three user credentials, and that ordering is
    load-bearing.

    A cookie is per-DOMAIN, so toup.ai holds exactly one `hex_sso_token`:
    whichever account signed in most recently. The bearer token and the
    `?token=` param are per-REQUEST and name the account the client
    actually means. Anyone with two accounts — a personal one and a
    demo/review one is the ordinary case — therefore has a browser where
    those two disagree, and the sidebar happily shows the localStorage
    identity while the cookie says someone else.

    With the cookie ranked second, such a request resolved to the WRONG
    user, the platform proxied to that user's agent, and the file was
    genuinely not there. The result is a fast, confident 404 on a file
    that exists — measured 2026-08-13 at 579 ms while the correct agent
    served the same file as a valid PDF. It reads as data loss.

    Explicit beats ambient: if the caller took the trouble to name a
    credential, honour it. The cookie stays as the fallback for embeds
    that carry nothing else.
    """
    user_id: Optional[str] = None

    # 1. Bearer header — the most explicit credential there is.
    auth_header = request.headers.get("authorization", "") if request else ""
    if auth_header.lower().startswith("bearer "):
        user_id = decode_access_token(auth_header[7:])

    # 2. Query param. Still explicit: the client put this account's token
    #    in the URL on purpose, so it outranks whatever cookie happens to
    #    be lying around for this domain.
    if not user_id and token:
        user_id = decode_access_token(token)

    # 3. SSO cookie — ambient, and only trusted when nothing was named.
    if not user_id and request is not None:
        cookie_tok = request.cookies.get(SSO_COOKIE_NAME)
        if cookie_tok:
            user_id = decode_access_token(cookie_tok)

    # Agent-mode fallback
    if not user_id and request is not None:
        agent_key = request.headers.get("x-agent-key", "")
        if settings.agent_api_key and agent_key == settings.agent_api_key and settings.user_id:
            user_id = settings.user_id

    if not user_id:
        raise HTTPException(status_code=401, detail="Authentication required")

    user = await get_user_by_id(db, user_id)
    if not user or not user.is_active:
        raise HTTPException(status_code=401, detail="User not found or inactive")
    return user


# Public name for the auth ladder above — the file library's download /
# preview routes need the same explicit-before-ambient order.
get_user_for_file = _get_user_for_file


async def _load_attachment(
    message_id: str, attachment_id: str, user_id: str, db: AsyncSession
) -> dict:
    """Fetch the attachment dict, verifying the caller owns the parent conversation.
    Returns 404 on miss OR on cross-user access (don't leak existence)."""
    msg = (await db.execute(
        select(Message).where(Message.id == message_id)
    )).scalar_one_or_none()
    if not msg or not msg.attachments:
        raise HTTPException(status_code=404, detail="Attachment not found")

    conv = (await db.execute(
        select(Conversation).where(Conversation.id == msg.conversation_id)
    )).scalar_one_or_none()
    if not conv or conv.user_id != user_id:
        raise HTTPException(status_code=404, detail="Attachment not found")

    # Column is JSON(B) — already a Python list when fetched via SQLAlchemy.
    # Belt-and-braces: some drivers may still return a string (legacy rows).
    attachments = msg.attachments
    if isinstance(attachments, str):
        try:
            attachments = json.loads(attachments)
        except (TypeError, ValueError):
            raise HTTPException(status_code=404, detail="Attachment not found")
    if not isinstance(attachments, list):
        raise HTTPException(status_code=404, detail="Attachment not found")

    for att in attachments:
        if isinstance(att, dict) and att.get("id") == attachment_id:
            return att
    raise HTTPException(status_code=404, detail="Attachment not found")


@router.get("/{message_id}/{attachment_id}")
async def download_file(
    message_id: str,
    attachment_id: str,
    request: Request,
    token: Optional[str] = Query(None),
    variant: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Stream the attachment bytes with correct Content-Type + Content-Disposition.

    `variant=thumb` serves the inline derivative written at persist time
    (`doc_generators.thumb_key`) — a WebP sized for the chat card, roughly a
    tenth of the bytes of a generated PNG. The chat card asks for it; the
    full-screen viewer and the download button never do, because the whole
    point of opening one is to inspect the real thing. A missing derivative
    falls through to the original rather than 404ing: it is an optimisation,
    and an image is worth delivering slowly.

    The response carries a strong ETag and `immutable` freshness, and a
    conditional GET is answered 304 without touching the file — the bytes
    behind one attachment id never change.
    """
    current_user = await _get_user_for_file(request, token, db)

    # Platform-side: data lives on the user's agent VPS. Proxy the request.
    if settings.run_mode == "platform":
        proxy = await _get_agent_proxy_info(current_user.id, db)
        if not proxy:
            raise HTTPException(status_code=404, detail="No active agent for user")
        content_iter, headers, status = await _proxy_file(
            proxy[0], proxy[1], f"{message_id}/{attachment_id}",
            {"variant": variant} if variant else {},
            conditional=_conditional_headers(request),
        )
        if status == 304:
            return Response(status_code=304, headers=headers)
        return StreamingResponse(content_iter, status_code=status,
                                 media_type=headers.get("content-type", "application/octet-stream"),
                                 headers=headers)

    att = await _load_attachment(message_id, attachment_id, current_user.id, db)
    key = att.get("storage_path")
    if not key:
        raise HTTPException(status_code=500, detail="Attachment has no storage_path")

    backend = get_storage_backend()
    if not backend.exists(key):
        logger.warning("Attachment file missing on disk: key=%s message=%s", key, message_id)
        raise HTTPException(status_code=410, detail="File has been deleted from storage")

    filename = att.get("filename", "download.bin")
    mime = att.get("mime_type", "application/octet-stream")

    # `served_thumb` tracks what we ACTUALLY serve, not what was asked for: a
    # request for the derivative that falls through to the original must not
    # be tagged as the derivative, or a client holding the thumbnail would be
    # told its copy of the full-size image is current.
    served_thumb = False
    if (variant or "").lower() == "thumb":
        from app.agent.doc_generators import thumb_key
        tkey = thumb_key(key)
        if backend.exists(tkey):
            key, mime, served_thumb = tkey, "image/webp", True

    size = backend.size(key)
    etag = attachment_etag(attachment_id, size, "thumb" if served_thumb else "")
    if etag_matches(request, etag):
        return not_modified(etag, _ATTACHMENT_CACHE)

    return StreamingResponse(
        stream_file(key),
        media_type=mime,
        headers={
            "Content-Length": str(size),
            "Content-Disposition": f'attachment; filename="{filename}"',
            # Original and derivative alike are content-addressed by the
            # attachment id and never change, so both may be cached hard.
            # Revisiting a thread then costs nothing at all — which is the
            # other half of the latency fix, and the half the original was
            # missing: at max-age=3600 with no validator, an image looked at
            # yesterday was re-downloaded in full today.
            "Cache-Control": _ATTACHMENT_CACHE,
            "ETag": etag,
        },
    )


@router.get("/{message_id}/{attachment_id}/preview")
async def preview_file(
    message_id: str,
    attachment_id: str,
    request: Request,
    format: str = Query("html"),
    token: Optional[str] = Query(None),
    db: AsyncSession = Depends(get_db),
):
    """Server-side render for the DocumentViewer pane.

    PDF / images → the file itself, `Content-Disposition: inline`, so an
                  <iframe>/WebView renders it natively. (Until 2026-08-18
                  these fell through to the 415 below even though
                  ws_chat.py / day_chats.py advertise a preview_url for
                  them — the iOS attachment card opened that URL in a
                  WebView and painted the JSON error body.)
    DOCX / PPTX → PDF via LibreOffice headless (cached). Browser renders
                  the PDF natively in an <iframe>, matching Slack /
                  Dropbox / Google Drive faithfulness.
    XLSX        → HTML tables (one per sheet) via openpyxl.
    Other       → 415. A caller that accepts text/html (browser, WebView
                  iframe) gets a small self-contained "Preview unavailable
                  · Download" page; everyone else gets a structured JSON
                  error with the download_url in it. Either way the client
                  should use /download.

    First DOCX/PPTX preview for a given file takes ~3-8s (LibreOffice
    cold start); subsequent requests are near-instant from the cache
    entry at `{storage_path}.preview.pdf`.
    """
    current_user = await _get_user_for_file(request, token, db)

    # Platform-side: proxy the preview request to the user's VPS agent.
    if settings.run_mode == "platform":
        proxy = await _get_agent_proxy_info(current_user.id, db)
        if not proxy:
            raise HTTPException(status_code=404, detail="No active agent for user")
        content_iter, headers, status = await _proxy_file(
            proxy[0], proxy[1], f"{message_id}/{attachment_id}/preview", {"format": format},
            accept=request.headers.get("accept") if request else None,
            read_timeout=_PREVIEW_READ_S,
            conditional=_conditional_headers(request),
        )
        if status == 304:
            return Response(status_code=304, headers=headers)
        return StreamingResponse(content_iter, status_code=status,
                                 media_type=headers.get("content-type", "text/html"),
                                 headers=headers)

    att = await _load_attachment(message_id, attachment_id, current_user.id, db)
    key = att.get("storage_path")
    mime = att.get("mime_type", "")
    backend = get_storage_backend()
    if not key or not backend.exists(key):
        raise HTTPException(status_code=410, detail="File unavailable")

    size = backend.size(key)
    return await render_preview(
        request,
        stream=lambda: stream_file(key),
        open_file=lambda: backend.open(key),
        path_of=lambda: backend.path(key),
        size=size,
        backend_key=key,
        mime=mime,
        filename=att.get("filename") or "document",
        download_url=f"{settings.api_prefix}/files/{message_id}/{attachment_id}",
        token=token,
        format=format,
        etag=attachment_etag(attachment_id, size),
        cache_control=_ATTACHMENT_CACHE,
    )


async def render_preview(
    request: Optional[Request],
    *,
    stream,
    size: int,
    mime: str,
    filename: str,
    download_url: str,
    open_file=None,
    path_of=None,
    token: Optional[str] = None,
    format: str = "html",
    backend_key: Optional[str] = None,
    etag: Optional[str] = None,
    cache_control: str = _DERIVED_CACHE,
) -> Response:
    """Server-side preview for one file — shared by the chat attachment route
    above and the file library (app/api/library.py).

    ``stream()`` returns an async iterator of bytes (``stream_file(key)`` for
    the storage backend, a plain chunked read for a library path): PDFs and
    images are served straight from it and never need a local path — an S3
    backend has none. ``open_file()`` (binary handle) is only used by the
    legacy mammoth DOCX path and ``path_of()`` only by XLSX → HTML, which
    needs a file on disk; either may be None, in which case those types
    get the download-only fallback. ``backend_key`` is the file_storage key
    (relative to generated/) when the file lives in the storage backend;
    the LibreOffice DOCX/PPTX → PDF path needs it for its cache entry.

    ``etag`` is the caller's validator for the SOURCE bytes; when given, a
    matching If-None-Match is answered 304 before any work happens — which
    for the DOCX/PPTX branch means before LibreOffice is asked for anything.
    Callers whose bytes can be replaced under a stable URL (the file library)
    must fold something that changes — size, mtime — into it, and leave
    ``cache_control`` at the revalidating default. Only chat attachments,
    whose id is minted per file, may pass `immutable`.

    The HTML branches (mammoth DOCX, XLSX tables) deliberately carry no
    validator: they are small, generated documents whose markup follows the
    server, and a 304 on one would pin a layout that shipped months ago.
    """
    mime = mime or "application/octet-stream"
    filename = filename or "document"

    # ── PDF / images → the bytes themselves, inline ───────────────
    # A PDF IS its own preview: every browser and WKWebView renders
    # application/pdf natively. Same for images. `inline` (not
    # `attachment`) is what makes an <iframe src=…> display rather than
    # download.
    if mime == "application/pdf" or mime.startswith("image/"):
        if etag and etag_matches(request, etag):
            return not_modified(etag, cache_control)
        headers = {
            "Content-Length": str(int(size or 0)),
            "Content-Disposition": f'inline; filename="{_ascii_filename(filename)}"',
            "Cache-Control": cache_control,
            # Nothing on this route is user-authored markup, but a
            # served image/PDF must never be sniffed into something
            # executable.
            "X-Content-Type-Options": "nosniff",
        }
        if etag:
            headers["ETag"] = etag
        return StreamingResponse(stream(), media_type=mime, headers=headers)

    # ── DOCX / PPTX → PDF via LibreOffice ────────────────────────
    # Production-grade faithful rendering. Cached per file.
    if backend_key and mime in (
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/vnd.openxmlformats-officedocument.presentationml.presentation",
    ):
        # Its own validator: the render is not the source, and a client
        # holding the PDF must not be told its copy of the DOCX is current.
        # Checked BEFORE render_to_pdf — the cheapest conversion measured is
        # 0.2 s from cache, and a 304 should cost nothing at all.
        pdf_etag = _suffix_etag(etag, "pdf")
        if pdf_etag and etag_matches(request, pdf_etag):
            return not_modified(pdf_etag, _DERIVED_CACHE)
        from app.services.doc_preview import render_to_pdf
        try:
            pdf_bytes = await render_to_pdf(backend_key)
        except RuntimeError as e:
            # Binary missing — agent needs libreoffice-{core,writer,impress}.
            raise HTTPException(status_code=501, detail=str(e))
        if not pdf_bytes:
            raise HTTPException(status_code=502, detail="LibreOffice conversion failed")
        headers = {
            "Cache-Control": _DERIVED_CACHE,
            "Content-Disposition": f'inline; filename="{_ascii_filename(filename)}.pdf"',
        }
        if pdf_etag:
            headers["ETag"] = pdf_etag
        return Response(content=pdf_bytes, media_type="application/pdf", headers=headers)

    if format != "html" and format != "png":
        raise HTTPException(status_code=400, detail="format must be 'html' or 'png'")

    # Legacy DOCX → mammoth HTML path is now dead (LibreOffice path above wins
    # first), but kept here for the rare case a deploy lacks libreoffice.
    if mime == "application/vnd.openxmlformats-officedocument.wordprocessingml.document":
        try:
            import mammoth  # type: ignore
        except ImportError:
            raise HTTPException(status_code=501, detail="mammoth not installed")
        if open_file is None:
            return _preview_unavailable(request, download_url, filename, mime, size, token)
        with open_file() as f:
            result = mammoth.convert_to_html(f)
        # Minimal shell so pane renders with basic typography.
        html = (
            "<!doctype html><meta charset='utf-8'>"
            "<style>body{font-family:system-ui,-apple-system,sans-serif;"
            "line-height:1.6;max-width:760px;margin:2rem auto;padding:0 1rem;color:#222}"
            "table{border-collapse:collapse}td,th{border:1px solid #ccc;padding:.3em .6em}"
            "img{max-width:100%}</style>"
            f"<body>{result.value}</body>"
        )
        return HTMLResponse(content=html)

    # ── XLSX → HTML (one <table> per sheet with tabs) ─────────────
    if mime == "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet":
        try:
            from openpyxl import load_workbook  # type: ignore
        except ImportError:
            raise HTTPException(status_code=501, detail="openpyxl not installed")
        import html as _html  # stored-XSS guard (audit-2026 re-audit round 11)
        if path_of is None:
            return _preview_unavailable(request, download_url, filename, mime, size, token)
        wb = load_workbook(path_of(), read_only=True, data_only=True)
        parts: list[str] = [
            "<!doctype html><meta charset='utf-8'>",
            "<style>body{font-family:system-ui,-apple-system,sans-serif;margin:0;padding:1rem;color:#222}"
            ".tabs{display:flex;gap:.25rem;margin-bottom:.5rem;border-bottom:1px solid #ddd}"
            ".tab{padding:.4rem .8rem;cursor:pointer;border:1px solid #ddd;border-bottom:none;"
            "background:#f6f6f6;border-radius:.3rem .3rem 0 0;user-select:none}"
            ".tab.active{background:#fff;font-weight:600}"
            ".sheet{display:none;overflow:auto;max-height:80vh}"
            ".sheet.active{display:block}"
            "table{border-collapse:collapse;font-size:.9em}"
            "td,th{border:1px solid #ddd;padding:.25em .5em;white-space:nowrap}"
            "th{background:#f6f6f6;font-weight:600}</style>",
            "<div class='tabs'>",
        ]
        for i, name in enumerate(wb.sheetnames):
            parts.append(
                f"<div class='tab{' active' if i == 0 else ''}' onclick=\"for(var e of document.querySelectorAll('.tab,.sheet'))e.classList.remove('active');this.classList.add('active');document.getElementById('s{i}').classList.add('active')\">{_html.escape(str(name))}</div>"
            )
        parts.append("</div>")
        for i, name in enumerate(wb.sheetnames):
            ws = wb[name]
            parts.append(f"<div class='sheet{' active' if i == 0 else ''}' id='s{i}'><table>")
            for row_idx, row in enumerate(ws.iter_rows(values_only=True)):
                tag = "th" if row_idx == 0 else "td"
                parts.append("<tr>" + "".join(
                    f"<{tag}>{'' if v is None else _html.escape(str(v))}</{tag}>" for v in row
                ) + "</tr>")
            parts.append("</table></div>")
        return HTMLResponse(content="".join(parts))

    # (PPTX is now handled by the LibreOffice PDF path above; legacy 204
    # fallback removed.)

    return _preview_unavailable(request, download_url, filename, mime, size, token)


def _ascii_filename(name: str) -> str:
    """A Content-Disposition-safe filename (quotes/newlines out, non-ASCII
    folded). The download routes send the full UTF-8 name via filename*."""
    cleaned = (name or "file").replace('"', "'").replace("\r", " ").replace("\n", " ")
    try:
        cleaned.encode("latin-1")
        return cleaned
    except UnicodeEncodeError:
        return cleaned.encode("ascii", "ignore").decode("ascii") or "file"


def _preview_unavailable(
    request: Optional[Request],
    download_url: str,
    filename: str,
    mime: str,
    size: int,
    token: Optional[str],
) -> Response:
    """415 for a type this endpoint cannot render — in the shape the caller
    can actually use.

    A browser / WebView navigation (Accept: text/html) gets a tiny
    self-contained page: filename, "Preview unavailable", a Download
    button. It carries the same `?token=` the caller used, because an
    <iframe>/WebView cannot send headers and that token is how it got
    here. An API caller (fetch/XHR, Accept */* or application/json) gets a
    structured JSON error with the download_url — no HTML in a JSON body,
    no JSON in a page.

    Before 2026-08-18 this was `HTTPException(415, "No preview for
    <mime>")` — fine for fetch(), but painted verbatim as
    `{"detail":"…"}` (and, via the platform proxy, double-encoded) by any
    client that rendered the response as a page.
    """
    import html as _html
    from urllib.parse import quote

    mime = mime or "application/octet-stream"
    filename = filename or "document"
    size = int(size or 0)
    accept = (request.headers.get("accept") or "").lower() if request else ""
    wants_html = "text/html" in accept

    if not wants_html:
        return JSONResponse(
            status_code=415,
            content={
                "detail": f"Preview unavailable for {mime}",
                "code": "preview_unavailable",
                "mime_type": mime,
                "filename": filename,
                "size_bytes": size,
                "download_url": download_url,
            },
        )

    href = download_url + (f"?token={quote(token)}" if token else "")
    page = (
        "<!doctype html><html><head><meta charset='utf-8'>"
        "<meta name='viewport' content='width=device-width,initial-scale=1'>"
        f"<title>{_html.escape(filename)}</title>"
        "<style>"
        "html,body{height:100%;margin:0}"
        "body{display:flex;align-items:center;justify-content:center;"
        "font-family:-apple-system,system-ui,'Segoe UI',Roboto,sans-serif;"
        "background:#0f0f12;color:#e6e6ea;text-align:center;padding:24px;box-sizing:border-box}"
        "@media (prefers-color-scheme: light){body{background:#f6f6f8;color:#1a1a1f}}"
        ".card{max-width:360px}"
        "h1{font-size:15px;font-weight:600;margin:0 0 6px;word-break:break-all}"
        "p{font-size:13px;opacity:.7;margin:0 0 18px}"
        "a{display:inline-block;padding:10px 18px;border-radius:10px;"
        "background:#6d5efc;color:#fff;text-decoration:none;font-size:14px;font-weight:600}"
        "</style></head><body><div class='card'>"
        f"<h1>{_html.escape(filename)}</h1>"
        f"<p>Preview unavailable · {_html.escape(mime)}"
        f"{' · ' + _fmt_bytes(size) if size else ''}</p>"
        f"<a href=\"{_html.escape(href, quote=True)}\" download=\"{_html.escape(filename, quote=True)}\">Download</a>"
        "</div></body></html>"
    )
    return HTMLResponse(
        content=page,
        status_code=415,
        headers={"Cache-Control": "no-store", "X-Content-Type-Options": "nosniff"},
    )


def _fmt_bytes(n: int) -> str:
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.1f} KB"
    return f"{n / (1024 * 1024):.1f} MB"
